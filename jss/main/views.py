#from django.http import HttpResponse
from django.shortcuts import render,redirect
from django.conf import settings
# from . import models
import json
# import openpyxl
from django.views.decorators.csrf import csrf_exempt
from main.models import Service,Parts
curl = settings.CURRENT_URL
#settings.USER_LOGIN
import datetime
#from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os
from openpyxl import Workbook ,load_workbook
date = datetime.datetime.now().date()
today = date.today()
d2 = str(today.strftime("%d-%m-%Y"))
@csrf_exempt
def Home(request):
    out = Parts.objects.all().order_by('pname')
    if request.method == "POST":
        cnames = request.POST.get('cname', '')
        cname=cnames.upper()
        mobile= request.POST.get('mobile', '')
        bnames = request.POST.get('bname', '')
        bname = bnames.upper()
        parts = request.POST.getlist('parts', '')
        amount = request.POST.get('amount', '')
        rnumbers = request.POST.get('rnumber', '')
        rnumber=rnumbers.upper()
        #print(cname,mobile,bname,parts,amount,rnumber)
        partslist=parts
        flag = "false"

        #Database value saving

        service = Service(cname=cname, mobile=mobile, bname=bname, parts=str(parts),amount=amount,rnumber=rnumber)
        if(Service.objects.filter(servicedate=datetime.datetime.now().date())):
            flag="true"
            #print(flag)
        else:
            # Stock related logic goes here
            for i in partslist:
                x = Parts.objects.get(pname=i)
                x.quantity = x.quantity - 1
                x.sold = x.sold + 1
                x.save() #updating stock
            service.save() #updating service


            #Excel writing started
            now = datetime.datetime.now()
            hr=now.hour
            mn=now.minute
            today = date.today()
            d1 = str(today.strftime("%d-%b"))+" "+str(hr)+":"+str(mn)
            filename = "Stock_Excel_File"
            directrie = "C:/Stock/"
            if not os.path.exists(directrie):
                os.makedirs(directrie)
            filedirect=str(directrie)+str(filename)+".xlsx"
            wb = load_workbook(filedirect)
            sheet = wb.active
            #------------------------------------------------------------
            sheet.column_dimensions['A'].width = 6
            sheet['A1'] = 'SNo.'
            sheet.column_dimensions['B'].width = 18
            sheet.column_dimensions['C'].width = 35
            sheet['C1'] = 'Customer Name'
            sheet.column_dimensions['D'].width = 20
            sheet['D1'] = 'Mobile'
            sheet.column_dimensions['E'].width = 18
            sheet['E1'] = 'Bike'
            sheet.column_dimensions['F'].width = 50
            sheet['F1'] = 'Parts Used'
            sheet.column_dimensions['G'].width = 20
            sheet['B1'] = 'Date'
            sheet['G1'] = 'Bike Number'
            sheet.column_dimensions['H'].width = 20
            sheet['H1'] = 'AMOUNT'
            parts=str(parts)
            var1=sheet.max_row-1
            partsvar = ''
            for part in parts:
                partsvar += '' + part
            list1=[var1,d1,cname,mobile,bname,partsvar,rnumber,amount]
            sheet.append(list1)
            wb.save(filedirect)
        return render(request, "index.html", {'curl': curl, 'out': out})
    else:
        return render(request, "index.html", {'curl': curl,'out':out})


def AddParts(request):
    flag="false"
    if request.method == "POST":
        pnames = request.POST.get('pname', '')
        pname=pnames.upper()
        mrp = request.POST.get('mrp', '')
        quantity = request.POST.get('quantity', '')
        places=request.POST.get('place','')
        place=places.upper()
        # print(pname,mrp,quantity,place)
        parts = Parts(pname=pname, MRP=mrp, quantity=quantity,place=place)
        if (Parts.objects.filter(pname=pname, MRP=mrp, quantity=quantity,place=place)):
            flag = "true"
            print(flag)
        else:
            parts.save()  # adding Parts
        return render(request, "addparts.html", {'curl': curl})
    else:
        return render(request,"addparts.html",{'curl': curl})

def login(request):
    if(request.method == "GET"):
        if (request.GET.get('action', '') == "login"):
            return render(request, "login.html", {'curl': curl,'msg':""})
    else:
        out=Parts.objects.values()
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        print(username,password)
        if (username == "HERO88178" and password == "29672967"):
            settings.USER_LOGIN = "1"
            out = Parts.objects.values()
            count = len(out)
            return render(request, "updateparts.html", {'curl': curl, 'out': out, 'count': count})
        else:
            return render(request, "login.html", {'curl': curl, 'msg': "INVALID CREDENTIAL"})


def Updateparts(request):
    if request.method == "POST":
        a=json.loads(request.POST.get('d', ''))
        l=len(a)-1
        delid = a['5000']
        for i in delid:
            Parts.objects.get(id=int(i)).delete()
        for i in range(1,l):
            j=str(i)
            print(a[j])
            id=int(a[j][0])#MRP,quantity,sold,place
            part = Parts.objects.get(id=id)
            part.pname=a[j][1].upper()
            part.MRP=float(a[j][3])
            part.quantity=int(a[j][2])
            part.sold=int(a[j][4])
            part.place=a[j][5].upper()
            part.save()
        return render(request, "success.html", {'curl': curl})

